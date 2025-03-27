import json
import calendar
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from rest_framework.response import Response
from rest_framework.decorators import api_view
from pymodbus.client import ModbusTcpClient
from pymodbus.exceptions import ModbusIOException
import logging
import pandas as pd
from django.shortcuts import render

from django.http import HttpResponse, JsonResponse

from miapp.energia.models import RegistroEnergia
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
# Configuración de conexión Modbus
MODBUS_HOST = "multimetersolin.dyndns.org"
MODBUS_PORT = 502
MODBUS_SLAVE_ID = 5

# Configuración del logger
logger = logging.getLogger(__name__)

def dashboard(request):
    """Renderiza la página principal del monitor de energía."""
    return render(request, 'dashboard.html')

def read_register(address, count=2, scale=1):
    """Función para leer registros Modbus con manejo de errores y logs."""
    client = ModbusTcpClient(MODBUS_HOST, port=MODBUS_PORT)
    client.connect()
    try:
        result = client.read_holding_registers(address=address, count=count, slave=MODBUS_SLAVE_ID)
        if result and not result.isError() and result.registers:
            raw_value = (result.registers[0] << 16) | result.registers[1]  # Convertir a 32 bits
            value = raw_value / scale
            logger.info(f"Lectura exitosa: {address} -> {value}")
            return value
        else:
            logger.warning(f"Error al leer dirección {address}")
            return "Error"
    except Exception as e:
        logger.error(f"Excepción en la lectura Modbus: {e}")
        return "Error"
    finally:
        client.close()

@api_view(['GET'])
def obtener_datos_modbus(request):
    """ Endpoint para obtener datos y guardarlos en la BD """
    datos = {
        "Voltaje Fase 1 (V)": read_register(4096, 2, 1000),
        "Voltaje Fase 2 (V)": read_register(4098, 2, 1000),
        "Voltaje Fase 3 (V)": read_register(4100, 2, 1000),
        "Corriente Fase 1 (A)": read_register(4102, 2, 1000),
        "Corriente Fase 2 (A)": read_register(4104, 2, 1000),
        "Corriente Fase 3 (A)": read_register(4106, 2, 1000),
        "Potencia Activa Total (kW)": read_register(4112, 2, 1000),
        "Potencia Reactiva Total (kVAR)": read_register(4114, 2, 1000),
        "Potencia Aparente Total (kVA)": read_register(4116, 2, 1000),
        "Energía Activa Total (kWh)": read_register(4688, 2, 10),
        "Energía Reactiva Total (kVARh)": read_register(4690, 2, 10),
    }

    # Guardar en la base de datos
    # Solo guardar si todos los datos requeridos son válidos (no "Error")
    if all(isinstance(v, (int, float)) for v in datos.values()):
        RegistroEnergia.objects.create(
            voltaje_fase_1=datos["Voltaje Fase 1 (V)"],
            voltaje_fase_2=datos["Voltaje Fase 2 (V)"],
            voltaje_fase_3=datos["Voltaje Fase 3 (V)"],
            corriente_fase_1=datos["Corriente Fase 1 (A)"],
            corriente_fase_2=datos["Corriente Fase 2 (A)"],
            corriente_fase_3=datos["Corriente Fase 3 (A)"],
            potencia_activa=datos["Potencia Activa Total (kW)"],
            potencia_reactiva=datos["Potencia Reactiva Total (kVAR)"],
            potencia_aparente=datos["Potencia Aparente Total (kVA)"],
            energia_activa=datos["Energía Activa Total (kWh)"],
            energia_reactiva=datos["Energía Reactiva Total (kVARh)"],
        )
    else:
        logger.warning("No se guardó el registro por valores inválidos.")

    return Response(datos)

def historico(request):
    return render(request, 'historico.html')


def obtener_historico(request):
    rango = request.GET.get('rango', 'diario')
    fecha_hoy = datetime.now()

    if rango == "diario":
        fecha_inicio = fecha_hoy - timedelta(days=1)
        fecha_fin = fecha_hoy
    elif rango == "semanal":
        fecha_inicio = fecha_hoy - timedelta(weeks=1)
        fecha_fin = fecha_hoy
    elif rango == "mensual":
        # Primer y último día del mes actual
        inicio_mes = datetime(fecha_hoy.year, fecha_hoy.month, 1)
        ultimo_dia = calendar.monthrange(fecha_hoy.year, fecha_hoy.month)[1]
        fin_mes = datetime(fecha_hoy.year, fecha_hoy.month, ultimo_dia, 23, 59, 59)
        fecha_inicio = inicio_mes
        fecha_fin = fin_mes
    else:
        fecha_inicio = fecha_hoy - timedelta(days=1)
        fecha_fin = fecha_hoy

    registros = RegistroEnergia.objects.filter(timestamp__range=(fecha_inicio, fecha_fin)).order_by('timestamp')

    data = [
        {
            "fecha": reg.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "voltaje_fase1": reg.voltaje_fase_1,
            "voltaje_fase2": reg.voltaje_fase_2,
            "voltaje_fase3": reg.voltaje_fase_3,
            "corriente_fase1": reg.corriente_fase_1,
            "corriente_fase2": reg.corriente_fase_2,
            "corriente_fase3": reg.corriente_fase_3,
            "potencia_activa": reg.potencia_activa,
            "potencia_reactiva": reg.potencia_reactiva,
            "potencia_aparente": reg.potencia_aparente,
            "energia_activa": reg.energia_activa,
            "energia_reactiva": reg.energia_reactiva
        }
        for reg in registros
    ]

    return JsonResponse(data, safe=False)


def exportar_excel(request):
    rango = request.GET.get('rango', 'diario')
    fecha_hoy = datetime.now()

    if rango == "diario":
        fecha_inicio = fecha_hoy - timedelta(days=1)
    elif rango == "semanal":
        fecha_inicio = fecha_hoy - timedelta(weeks=1)
    elif rango == "mensual":
        fecha_inicio = fecha_hoy - timedelta(days=30)
    else:
        fecha_inicio = fecha_hoy - timedelta(days=1)

    registros = RegistroEnergia.objects.filter(timestamp__gte=fecha_inicio).order_by('timestamp')

    df = pd.DataFrame([{
        "Fecha": r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        "Voltaje Fase 1 (V)": r.voltaje_fase_1,
        "Voltaje Fase 2 (V)": r.voltaje_fase_2,
        "Voltaje Fase 3 (V)": r.voltaje_fase_3,
        "Corriente Fase 1 (A)": r.corriente_fase_1,
        "Corriente Fase 2 (A)": r.corriente_fase_2,
        "Corriente Fase 3 (A)": r.corriente_fase_3,
        "Energía Activa Total (kWh)": r.energia_activa,
        "Energía Reactiva Total (kVARh)": r.energia_reactiva
    } for r in registros])

    # Crear libro de trabajo
    wb = Workbook()
    resumen_ws = wb.active
    resumen_ws.title = "Resumen"
    historico_ws = wb.create_sheet("Histórico de Datos")

    # --- Hoja de Resumen ---
    resumen_ws.merge_cells("A1:E1")
    resumen_ws["A1"] = "🔌 Resumen de Consumo Energético"
    resumen_ws["A1"].font = Font(size=14, bold=True)
    resumen_ws["A1"].alignment = Alignment(horizontal="center")

    def promedio(col):
        return df[col].dropna().mean()

    def ultimo(col):
        return df[col].dropna().iloc[-1] if not df[col].dropna().empty else 0

    resumen_ws.append(["Promedio Voltaje Fase 1 (V)", promedio("Voltaje Fase 1 (V)")])
    resumen_ws.append(["Promedio Voltaje Fase 2 (V)", promedio("Voltaje Fase 2 (V)")])
    resumen_ws.append(["Promedio Voltaje Fase 3 (V)", promedio("Voltaje Fase 3 (V)")])
    resumen_ws.append([])
    resumen_ws.append(["Última Energía Activa Total (kWh)", ultimo("Energía Activa Total (kWh)")])
    resumen_ws.append(["Última Energía Reactiva Total (kVARh)", ultimo("Energía Reactiva Total (kVARh)")])

    # Estilo para toda la hoja
    for row in resumen_ws.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")

    # --- Hoja de Histórico de Datos ---
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        historico_ws.append(row)
        if r_idx == 1:
            for cell in historico_ws[r_idx]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="d1ecf1", fill_type="solid")

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="resumen_energia.xlsx"'
    wb.save(response)
    return response